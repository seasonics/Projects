from __future__ import division # for division
import sys
import codecs
import re
import time
import datetime
import os
import binascii
import string
import cPickle as pickle
import json
from DBConnection import MysqlConnecttion
from openpyxl import load_workbook
from booth.extractor import Extractor
import csv

class Labels:
        def __init__(self):
            self.word_label_map = {}
        def init_label_map(self):
                path = r"/Users/elianapolimeni/Desktop/Cole_booth_work/Labels.xlsx"
                sql = MysqlConnecttion()
                wb = load_workbook(path)
                sheet = wb.get_sheet_by_name('Labels.csv');
                row = 1
                word = ""
                keep = 0
                size = len(sheet.rows)
                while row < size:
                        label_id = sheet.cell(row=row, column=0).value
                        removed = sheet.cell(row=row, column=1).value
                        word = sheet.cell(row=row, column=2).value
                        natural_label = sheet.cell(row=row, column=3).value
                        short_label = sheet.cell(row=row, column=5).value
                        geography = sheet.cell(row=row, column = 6).value
                        vertical = sheet.cell(row=row, column = 7).value
                        market_segment = sheet.cell(row=row, column = 8).value
                        if removed == "yes":
                                row+=1;
                                continue
                        if geography == None or geography == "":
                                geography = 0
                        if vertical == None or vertical == "":
                                vertical = 0
                        if market_segment == None or market_segment == "":
                                market_segment = 0
                        if natural_label == None:
                            natural_label = ""
                        if short_label == None:
                            short_label = ""
                        map = {}
                        map["id"] = label_id
                        map["natural_label"] = natural_label
                        map["short_label"] = short_label
                        map["geogr"] = geography
                        map["vert"] = vertical
                        map["mktseg"] = market_segment
                        self.word_label_map[word] = map
                        row+=1;
        def fill_label_table(self):
                """
                put the word/phrase into database
                """
                local_sql = MysqlConnecttion("local")
                #booth_sql = MysqlConnecttion("booth")
                self.init_label_map()
                print len(self.word_label_map)
                for word in self.word_label_map:
                        map = self.word_label_map[word]
                        query = r'''insert into Labels (Labelid, Word, Natural_label, Short_label, Geogr, Vert, Mktseg) values
                                            ('%s', '%s', '%s', '%s', '%s', '%s', '%s')
                                            on duplicate key update
                                            Word = '%s',
                                            Natural_label = '%s', 
                                            Short_label = '%s',
                                            Geogr = '%s', 
                                            Vert = '%s',
                                            Mktseg ='%s'
                                            '''%(
                                                map["id"], word, map["natural_label"], map["short_label"], map["geogr"], map["vert"], map["mktseg"],
                                                word, map["natural_label"],map["short_label"],map["geogr"],map["vert"],map["mktseg"])
                        # print query
                        local_sql.excute(query)
                        #booth_sql.excute(query)
        def update_super_label(self):
                """
                update the labels table to fill the super_labid
                """
                path = r"C:\Users\sun\Dropbox\Gartner Files\Jilong\network\Higher_level_labels (super).xlsx"
                sql = MysqlConnecttion()
                wb = load_workbook(path)
                sheet = wb.get_sheet_by_name('Sheet1');
                row_no = 1
                word = ""
                size = len(sheet.rows)
                label_map = {}
                while row_no < size:
                        word = sheet.cell(row=row_no, column=0).value
                        query = r"select Labelid from labels where Word = '%s'" % (word)
                        rows = sql.excute_with_result(query)
                        row = rows[0]
                        label_map[word] = row[0]
                        row_no += 1
                for word in label_map:
                        id = label_map[word]
                        query = r"UPDATE labels SET Super_labid = %d WHERE word LIKE '%%%s%%' AND word != '%s'"%(id, word, word)
                        sql.excute(query)
        def update_mq_id(self):
                """
                1.  Numerical magic quadrant id:
                What I did was used the original id (file name), x 10 and then added 1, 2, 3, etc. for the a, b, and c's.
                So mq9a translates to 91, mq42c translates to 423.
                """
                sql = MysqlConnecttion()
                mq_map = {}
                query = r"select MQID from Magic_Quadrants where removed = 0"
                rows = sql.excute_with_result(query)
                for row in rows:
                        mq_map[row[0]] = 0
                pattern = re.compile(r"mq(\d+)(\w+)")
                for mq_id in mq_map:
                        match = pattern.match(mq_id)
                        base = int(match.group(1,)) * 10
                        digit = match.group(2,)
                        if digit == 'a':
                                base += 1
                        elif digit == 'b':
                                base += 2
                        elif digit == 'c':
                                base += 3
                        elif digit == 'd':
                                base += 4
                        elif digit == 'e':
                                base += 5
                        elif digit == 'f':
                                base += 6
                        elif digit == 'g':
                                base += 7
                        elif digit == 'h':
                                base += 8
                        else:
                                print mq_id
                        mq_map[mq_id] = base
                for mq_id in mq_map:
                        print mq_id
                        query = r"update Magic_Quadrants set mqid2 = '%d' where MQID = '%s'"%(mq_map[mq_id], mq_id)
                        sql.excute(query)
                for mq_id in mq_map:
                        print mq_id
                        query = r"update MQ_DETAIL set mqid2 = '%d' where MQID = '%s'"%(mq_map[mq_id], mq_id)
                        sql.excute(query)
        def update_flag_in_mq_table(self):
                """
                update the flag in magic_quadrants table according to their title
                """
                local_sql = MysqlConnecttion("local")
                booth_sql = MysqlConnecttion("booth")
                label_map = {}#{"word":labelid}
                query = r"select Word, Geogr, Vert, Mktseg from labels WHERE Geogr = 1 or Vert = 1 or Mktseg = 1"
                rows = local_sql.excute_with_result(query)
                for row in rows:
                        word_flag = {}
                        word_flag["geogr"] = int(row[1])
                        word_flag["vert"] = int(row[2])
                        word_flag["mktseg"] = int(row[3])
                        label_map[row[0]] = word_flag
                query = r'''(SELECT mqid, mqid2, mq_title_vector_short FROM magic_quadrants WHERE mqid NOT IN (
                                SELECT mqid FROM magic_quadrants WHERE mq_title_vector_unique IS NOT NULL AND mq_title_vector_unique != '') AND removed = 0)
                                UNION
                                (SELECT mqid, mqid2, mq_title_vector_unique FROM    magic_quadrants WHERE mq_title_vector_unique IS NOT NULL AND mq_title_vector_unique != ''   AND removed = 0)'''
                mq_vector_map = {}#{"mqid":"word vector (short)"}
                mqid_map = {}#{"mqid":mqid2}
                rows = local_sql.excute_with_result(query)
                for row in rows:
                        mq_vector_map[row[0]] = row[2]
                        mqid_map[row[0]] = row[1]
                for mq_id in mq_vector_map:
                        print mq_id
                        mq_flag = {"geogr":0, "vert":0, "mktseg":0}
                        json_word_set = mq_vector_map[mq_id]
                        if json_word_set == None or json_word_set == "":
                                json_word_set = "{}"
                        word_map = json.loads(json_word_set)
                        label_list = ""
                        for word in word_map:
                                if word in label_map:
                                        word_flag = label_map[word]
                                        mq_flag["geogr"] |= word_flag["geogr"]
                                        mq_flag["vert"] |= word_flag["vert"]
                                        mq_flag["mktseg"] |= word_flag["mktseg"]
                        query = r'''UPDATE Magic_Quadrants set Geo_label = '%s', Vert_label = '%s', Mkts_label = '%s' WHERE MQID = '%s'
                                        and removed = 0'''%(mq_flag["geogr"], mq_flag["vert"], mq_flag["mktseg"], mq_id)
                        # print query
                        local_sql.excute(query)
                        booth_sql.excute(query)
        def update_labelid(self):
                    """
                    update the labelid in Magic_Quadrants
                    """
                    local_sql = MysqlConnecttion("local")
                    # booth_sql = MysqlConnecttion("booth")
                    label_map = {}#{"word":labelid}
                    query = r"select Word, Labelid from labels"
                    rows = local_sql.excute_with_result(query)
                    for row in rows:
                        label_map[row[0]] = row[1]
                    query = r'''(SELECT mqid, mqid2, labelid, mq_title_vector_short FROM magic_quadrants where removed = 0)'''
                    mq_vector_map = {}#{"mqid":"word vector (short)"}
                    label_tmap = {}
                    rows = local_sql.excute_with_result(query)
                    for row in rows:
                        mq_vector_map[row[0]] = row[3]
                        label_map[row[0]] = row[2]   
                    for mq_id in mq_vector_map:
                        json_word_set = mq_vector_map[mq_id]
                        if json_word_set == None or json_word_set == "":
                                        json_word_set = "{}"
                        word_map = json.loads(json_word_set)
                        label_list = []
                        for word in word_map:
                            if word in label_map:    
                                label_list.append(str(label_map[word]))
                        labels = ";".join(label_list)
                        query = r"UPDATE Magic_Quadranta set Labelid = '' WHERE MQID = '%s' and removed = 0"%(mq_id)
                        query = r"UPDATE Magic_Quadrants set Labelid = '%s' WHERE MQID = '%s' and removed = 0"%(labels, mq_id)
                        local_sql.excute(query)
                        # booth_sql.excute(query)
                            
        def update_labelid_first(self):
            """
                update the labelid in Magic_Quadrants
                """
            local_sql = MysqlConnecttion("local")
            booth_sql = MysqlConnecttion("booth")
            label_map = {}#{"word":labelid}
            query = r"select Word, Labelid from labels WHERE Geogr = 0 and Vert = 0 and Mktseg = 0"
            rows = local_sql.excute_with_result(query)
            for row in rows:
                label_map[row[0]] = row[1]
            query = r'''(SELECT mqid, docid, labelid, mq_title_vector_short FROM magic_quadrants where removed = 0)'''
            mq_vector_map = {}#{"mqid":"word vector (short)"}
            label_tmap = {}
            docid_map ={}
            rows = local_sql.excute_with_result(query)
            for row in rows:
                mq_vector_map[row[0]] = row[3]
                docid_map[row[0]] = row[1]
                label_map[row[0]] = row[2]
            query = r'''(SELECT title_short, docid FROM doc_deatail_vector)'''
            cool_map ={}
            rows = local_sql.excute_with_result(query)
            for row in rows:
                cool_map[row[1]] = row[0]
            for mq_id in mq_vector_map:
                json_word_set = mq_vector_map[mq_id]
                if json_word_set == None or json_word_set == "":
                    json_word_set = "{}"
                word_map = json.loads(json_word_set)
                label_list = []
                for word in word_map:
                    if word in label_map:
                        label_list.append(str(label_map[word]))
                json_word_set = cool_map[docid_map[mq_id]]
                if json_word_set == None or json_word_set == "":
                    json_word_set = "{}"
                word_map = json.loads(json_word_set)
                for word in word_map:
                    if word in label_map:
                        label_list.append(str(label_map[word]))
                label_list = list(set(label_list))
                length = len(label_list)
                if length == 0:
                    query = r'''insert into new_magic_quadrants (DocID, MQID)
                        values
                        ('%s', '%s')
                        '''%(docid_map[mq_id], mq_id)
                    local_sql.excute(query)
                    booth_sql.excute(query)
                if length == 1:
                    query = r'''insert into new_magic_quadrants (DocID, MQID, Labelid1)
                        values
                        ('%s', '%s', '%s')
                        '''%(docid_map[mq_id], mq_id, label_list[0])
                    local_sql.excute(query)
                    booth_sql.excute(query)
                if length == 2:
                    query = r'''insert into new_magic_quadrants (DocID, MQID, Labelid1,Labelid2)
                        values
                        ('%s', '%s', '%s', '%s')
                        '''%(docid_map[mq_id], mq_id, label_list[0],label_list[1])
                    local_sql.excute(query)
                    booth_sql.excute(query)
                if length == 3:
                    query = r'''insert into new_magic_quadrants (DocID, MQID, Labelid1,Labelid2,Labelid3)
                        values
                        ('%s', '%s', '%s', '%s', '%s')
                        '''%(docid_map[mq_id], mq_id, label_list[0],label_list[1],label_list[2])
                    local_sql.excute(query)
                    booth_sql.excute(query)
                if length == 4:
                    query = r'''insert into new_magic_quadrants (DocID, MQID, Labelid1,Labelid2,Labelid3,Lableid4)
                        values
                        ('%s', '%s', '%s', '%s', '%s', '%s')
                        '''%(docid_map[mq_id], mq_id, label_list[0],label_list[1],label_list[2],label_list[3])
                    local_sql.excute(query)
                    booth_sql.excute(query)
                if length == 5:
                    query = r'''insert into new_magic_quadrants (DocID, MQID, Labelid1,Labelid2,Labelid3,Lableid4,labelid5)
                        values
                        ('%s', '%s', '%s', '%s', '%s', '%s', '%s')
                        '''%(docid_map[mq_id], mq_id, label_list[0],label_list[1],label_list[2],label_list[3],label_list[4])
                    local_sql.excute(query)
                    booth_sql.excute(query)
                if length > 5:
                    print "lolz"
                    return

        def check_hr(self):
            local_sql = MysqlConnecttion("local")
            booth_sql = MysqlConnecttion("booth")
            query = r"select MQID from new_magic_quadrants WHERE Labelid1 = '120' and Vert_label = '99'"
            rows = local_sql.excute_with_result(query)
            for row in rows:
                query = r"update new_magic_quadrants set Labelid1 = NULL where MQID = '%s'"%(row)
                local_sql.excute(query)
                booth_sql.excute(query)

        def update_labelid_geo(self):
            """
                update the labelid in Magic_Quadrants
                """
            local_sql = MysqlConnecttion("local")
            booth_sql = MysqlConnecttion("booth")
            label_map = {}#{"word":labelid}
            query = r"select Word, Labelid from labels WHERE Geogr = 1 and Vert = 0 and Mktseg = 0"
            rows = local_sql.excute_with_result(query)
            for row in rows:
                label_map[row[0]] = row[1]
            query = r'''(SELECT mqid, docid, labelid, mq_title_vector_short FROM magic_quadrants where removed = 0)'''
            mq_vector_map = {}#{"mqid":"word vector (short)"}
            label_tmap = {}
            docid_map ={}
            rows = local_sql.excute_with_result(query)
            for row in rows:
                mq_vector_map[row[0]] = row[3]
                docid_map[row[0]] = row[1]
                label_map[row[0]] = row[2]
            query = r'''(SELECT title_short, docid FROM doc_deatail_vector)'''
            cool_map ={}
            rows = local_sql.excute_with_result(query)
            for row in rows:
                cool_map[row[1]] = row[0]
            for mq_id in mq_vector_map:
                json_word_set = mq_vector_map[mq_id]
                if json_word_set == None or json_word_set == "":
                    json_word_set = "{}"
                word_map = json.loads(json_word_set)
                label_list = []
                for word in word_map:
                    if word in label_map:
                        label_list.append(str(label_map[word]))
                json_word_set = cool_map[docid_map[mq_id]]
                if json_word_set == None or json_word_set == "":
                    json_word_set = "{}"
                word_map = json.loads(json_word_set)
                for word in word_map:
                    if word in label_map:
                        label_list.append(str(label_map[word]))
                label_list = list(set(label_list))
                length = len(label_list)
                labels = ";".join(label_list)
                query = r"update new_magic_quadrants set Geo_label = '%s' where MQID = '%s'"%(labels, mq_id)
                local_sql.excute(query)
                booth_sql.excute(query)

        def update_labelid_vert(self):
            """
                update the labelid in Magic_Quadrants
                """
            local_sql = MysqlConnecttion("local")
            booth_sql = MysqlConnecttion("booth")
            label_map = {}#{"word":labelid}
            query = r"select Word, Labelid from labels WHERE Geogr = 0 and Vert = 1 and Mktseg = 0"
            rows = local_sql.excute_with_result(query)
            for row in rows:
                label_map[row[0]] = row[1]
            query = r'''(SELECT mqid, docid, labelid, mq_title_vector_short FROM magic_quadrants where removed = 0)'''
            mq_vector_map = {}#{"mqid":"word vector (short)"}
            label_tmap = {}
            docid_map ={}
            rows = local_sql.excute_with_result(query)
            for row in rows:
                mq_vector_map[row[0]] = row[3]
                docid_map[row[0]] = row[1]
                label_map[row[0]] = row[2]
            query = r'''(SELECT title_short, docid FROM doc_deatail_vector)'''
            cool_map ={}
            rows = local_sql.excute_with_result(query)
            for row in rows:
                cool_map[row[1]] = row[0]
            for mq_id in mq_vector_map:
                json_word_set = mq_vector_map[mq_id]
                if json_word_set == None or json_word_set == "":
                    json_word_set = "{}"
                word_map = json.loads(json_word_set)
                label_list = []
                for word in word_map:
                    if word in label_map:
                        label_list.append(str(label_map[word]))
                json_word_set = cool_map[docid_map[mq_id]]
                if json_word_set == None or json_word_set == "":
                    json_word_set = "{}"
                word_map = json.loads(json_word_set)
                for word in word_map:
                    if word in label_map:
                        label_list.append(str(label_map[word]))
                label_list = list(set(label_list))
                length = len(label_list)
                labels = ";".join(label_list)
                query = r"update new_magic_quadrants set Vert_label = '%s' where MQID = '%s'"%(labels, mq_id)
                local_sql.excute(query)
                booth_sql.excute(query)

        def update_labelid_intext(self):
            """
                update the labelid in Magic_Quadrants
                """
            local_sql = MysqlConnecttion("local")
            booth_sql = MysqlConnecttion("booth")
            label_map = {}#{"word":labelid}
            query = r"select Word, Labelid from labels WHERE Geogr = 0 and Vert = 0 and Mktseg = 0"
            rows = local_sql.excute_with_result(query)
            for row in rows:
                label_map[row[0]] = row[1]
            query = r'''(SELECT mqid, docid, labelid, mq_title_vector_short FROM magic_quadrants where removed = 0)'''
            mq_vector_map = {}#{"mqid":"word vector (short)"}
            label_tmap = {}
            docid_map ={}
            rows = local_sql.excute_with_result(query)
            for row in rows:
                mq_vector_map[row[0]] = row[3]
                docid_map[row[0]] = row[1]
                label_map[row[0]] = row[2]
            query = r'''(SELECT first_short,what_short,market_short, docid FROM doc_deatail_vector)'''
            first_map ={}
            what_map = {}
            market_map = {}
            rows = local_sql.excute_with_result(query)
            for row in rows:
                first_map[row[3]] = row[0]
                what_map[row[3]] = row[1]
                market_map[row[3]] = row[2]
            for mq_id in mq_vector_map:
                label_list =[]
                json_word_set = first_map[docid_map[mq_id]]
                if json_word_set == None or json_word_set == "":
                    json_word_set = "{}"
                word_map = json.loads(json_word_set)
                for word in word_map:
                    if word in label_map:
                        label_list.append(str(label_map[word]))
                json_word_set = what_map[docid_map[mq_id]]
                if json_word_set == None or json_word_set == "":
                    json_word_set = "{}"
                word_map = json.loads(json_word_set)
                for word in word_map:
                    if word in label_map:
                        label_list.append(str(label_map[word]))
                json_word_set = market_map[docid_map[mq_id]]
                if json_word_set == None or json_word_set == "":
                    json_word_set = "{}"
                word_map = json.loads(json_word_set)
                for word in word_map:
                    if word in label_map:
                        label_list.append(str(label_map[word]))
    
                label_list = list(set(label_list))
                labels = ";".join(label_list)
                query = r"update new_magic_quadrants set Labelid_intext = '%s' where MQID = '%s'"%(labels, mq_id)
                local_sql.excute(query)
                booth_sql.excute(query)

        def update_labelid_mkt(self):
            """
                update the labelid in Magic_Quadrants
                """
            local_sql = MysqlConnecttion("local")
            booth_sql = MysqlConnecttion("booth")
            label_map = {}#{"word":labelid}
            query = r"select Word, Labelid from labels WHERE Geogr = 0 and Vert = 0 and Mktseg = 1"
            rows = local_sql.excute_with_result(query)
            for row in rows:
                label_map[row[0]] = row[1]
            query = r'''(SELECT mqid, docid, labelid, mq_title_vector_short FROM magic_quadrants where removed = 0)'''
            mq_vector_map = {}#{"mqid":"word vector (short)"}
            label_tmap = {}
            docid_map ={}
            rows = local_sql.excute_with_result(query)
            for row in rows:
                mq_vector_map[row[0]] = row[3]
                docid_map[row[0]] = row[1]
                label_map[row[0]] = row[2]
            query = r'''(SELECT title_short, docid FROM doc_deatail_vector)'''
            cool_map ={}
            rows = local_sql.excute_with_result(query)
            for row in rows:
                cool_map[row[1]] = row[0]
            for mq_id in mq_vector_map:
                json_word_set = mq_vector_map[mq_id]
                if json_word_set == None or json_word_set == "":
                    json_word_set = "{}"
                word_map = json.loads(json_word_set)
                label_list = []
                for word in word_map:
                    if word in label_map:
                        label_list.append(str(label_map[word]))
                json_word_set = cool_map[docid_map[mq_id]]
                if json_word_set == None or json_word_set == "":
                    json_word_set = "{}"
                word_map = json.loads(json_word_set)
                for word in word_map:
                    if word in label_map:
                        label_list.append(str(label_map[word]))
                label_list = list(set(label_list))
                length = len(label_list)
                labels = ";".join(label_list)
                query = r"update new_magic_quadrants set Mkt_label = '%s' where MQID = '%s'"%(labels, mq_id)
                local_sql.excute(query)
                booth_sql.excute(query)

        def mq_title_shorten_word_vector(self):
            """
            shorten the word vector from the vector we already got
            """
            field = "mq_title_vector"
            local_sql = MysqlConnecttion("local")
            doc_freq = set()
            query = r"select word from labels"
            rows = local_sql.excute_with_result(query)
            for row in rows:
                doc_freq.add(row[0])
            query = r'''SELECT MQID, %s FROM magic_quadrants'''%(field)
            rows = local_sql.excute_with_result(query)
            for row in rows:
                docid = row[0]
                json_words = row[1]
                if(json_words == None or len(json_words) == 0):
                    continue
                word_vector = json.loads(json_words)
                result = sorted(word_vector.items(), key=lambda d:(d[0].count("_")),reverse = True)
                short_vector = {}
                for tuple in result:
                        key = tuple[0]
                        value = tuple[1]
                        if key in doc_freq:
                                is_in = 0
                                for word in short_vector:
                                        if(word.count(key) > 0):
                                                is_in += 1
                                if is_in == 0:
                                       short_vector[key] = value
                try:
                        if len(short_vector) > 0:
                                json_words = json.dumps(short_vector)
                                json_words = json_words.replace(r"'",r"\'")
                        else :
                                json_words = ""
                except UnicodeDecodeError, e:
                        print key
                        print short_vector
                        json_words = ''
                        raw_input()

                query = r"update magic_quadrants set %s_short = '%s' where MQID = '%s'"%(field, json_words, docid)
               
                local_sql.excute(query)
                
        def get_word_vector(self, type):
            """
            extract the word vector from First_Paragraph, What_you_need_to_know and MarketDefinition
            """
            if type == "First_Paragraph":
                field = "first"
            elif type == "What_you_need_to_know":
                field = "what"
            elif type == "MarketDefinition":
                field = "market"
            else:
                field = ""
            e =Extractor()
            local_sql = MysqlConnecttion("local")
            query = r"select distinct docid from Magic_Quadrants where removed = 0"
            rows = local_sql.excute_with_result(query)
            mq_set = set()
            for row in rows:
                mq_set.add(row[0])
            query = r"select count(*) from doc_deatail_vector"
            count = local_sql.excute_with_result(query)[0][0]
            if count == 0:
                for docid in mq_set:
                    query = r"insert into doc_deatail_vector set docid = '%s'"%docid
                    local_sql.excute(query)
            for docid in mq_set:
                print docid, type
                query = "select %s from doc_detail WHERE docid = '%s'"%(type, docid)
                row = local_sql.excute_with_result(query)[0]
                content = row[0]
                if content == None or len(content) == 0:
                    continue
                word_vector = e.extract(content, shorten = True)
                try:
                        json_words = json.dumps(word_vector)
                        json_words = json_words.replace(r"'",r"\'")
                except UnicodeDecodeError, e:
                        print key
                        print pair_with_weight
                        json_words = ''
                        raw_input()
                query = r"UPDATE doc_deatail_vector set %s = '%s' where docid = '%s'"%(field, json_words, docid)
                local_sql.excute(query)

        def shorten_word_vector(self, type):
            """
            shorten the word vector from the vector we already got
            """
            if type == "First_Paragraph":
                field = "first"
            elif type == "What_you_need_to_know":
                field = "what"
            elif type == "MarketDefinition":
                field = "market"
            else:
                field = ""
            local_sql = MysqlConnecttion("local")
            doc_freq = set()
            query = r"select word from labels"
            rows = local_sql.excute_with_result(query)
            for row in rows:
                doc_freq.add(row[0])
            query = r'''SELECT DocID, %s FROM doc_deatail_vector'''%(field)
            rows = local_sql.excute_with_result(query)
            for row in rows:
                docid = row[0]
                json_words = row[1]
                print docid, type
                if(json_words == None or len(json_words) == 0):
                    continue
                word_vector = json.loads(json_words)
                result = sorted(word_vector.items(), key=lambda d:(d[0].count("_")),reverse = True)
                short_vector = {}
                for tuple in result:
                        key = tuple[0]
                        value = tuple[1]
                        if(key in doc_freq):
                                is_in = 0
                                for word in short_vector:
                                        if(word.count(key) > 0):
                                                is_in += 1
                                if is_in == 0:
                                        short_vector[key] = value
                try:
                        if len(short_vector) > 0:
                                json_words = json.dumps(short_vector)
                                json_words = json_words.replace(r"'",r"\'")
                        else :
                                json_words = ""
                except UnicodeDecodeError, e:
                        print key
                        print short_vector
                        json_words = ''
                        raw_input()

                query = r"update doc_deatail_vector set %s_short = '%s' where DocID = '%s'"%(field, json_words, docid)
                #print query
                local_sql.excute(query)
        def mark_text_labels(self):
            """
            mark the text label
            """
            local_sql = MysqlConnecttion("local")
            booth_sql = MysqlConnecttion("booth")
            label_map = {}#{"word":labelid}
            query = r"select Word, Labelid from labels WHERE Geogr = 0 and Vert = 0 and Mktseg = 0"
            rows = local_sql.excute_with_result(query)
            for row in rows:
                    label_map[row[0]] = row[1]
            doc_map = {}
            query = r'''SELECT docid, first, what, market from doc_deatail_vector
                                '''
            rows = local_sql.excute_with_result(query)
            for row in rows:
                docid = row[0]
                print docid
                title_label = set()
                query = r"select Labelid from Magic_Quadrants where docid = '%s' and removed = 0"%docid
                label_rows = local_sql.excute_with_result(query)
                for r in label_rows:
                    labels = r[0].split(";")
                    for label in labels:
                        title_label.add(label)
                word_vector = set()
                json_words = row[1]
                if(json_words == None or len(json_words) == 0):
                    pass
                else:
                    first_vector = json.loads(json_words)
                    word_vector =word_vector.union(first_vector.keys())
                json_words = row[2]
                if(json_words == None or len(json_words) == 0):
                    pass
                else:
                    what_vector = json.loads(json_words)
                    word_vector =word_vector.union(what_vector.keys())
                json_words = row[3]
                if(json_words == None or len(json_words) == 0):
                    pass
                else:
                    market_vector = json.loads(json_words)
                    word_vector =word_vector.union(market_vector.keys())
                label_list = set()
                for word in word_vector:
                    if word in label_map and str(label_map[word]) not in title_label:
                        if str(label_map[word]) not in label_list:
                            label_list.add(str(label_map[word]))
                            query = r"insert into mq_text_label (DocID, label_id) values ('%s','%s')"%(docid, label_map[word])
                            # local_sql.excute(query)
                            # booth_sql.excute(query)
                text_labels = ";".join(label_list)
                query = r"update Magic_Quadrants set Labelid_intext = '%s' where DocID = '%s' and removed = 0"%(text_labels, docid)
                # print query
                local_sql.excute(query)
                # booth_sql.excute(query)

        def text_labels(self):
            """
            Add label id intext - these are labels that are included in the text
            of the doc but not in the title (text includes first par, what you need to know,
            and market def - DO NOT include key findings or key assumptions)
            1. need to extract the noun, verb, etc. for each
            2. get the short word vector for each
            3. the labels from each short vector
            """
            types = [
            "First_Paragraph",
            "What_you_need_to_know",
            "MarketDefinition"
            ]
            for t in types:
                # self.get_word_vector(t)
                self.shorten_word_vector(t)
            self.update_labelid_intext()
            return
            local_sql = MysqlConnecttion("local")
            # booth_sql = MysqlConnecttion("booth")
            query = r"select distinct docid from Magic_Quadrants where removed = 0"
            rows = local_sql.excute_with_result(query)
            mq_set = set()
            for row in rows:
                mq_set.add(row[0])
            doc_map = {}
            query = "select DocID, DocNo from Documents"
            rows = local_sql.excute_with_result(query)
            for row in rows:
                docid = row[0]
                docno = row[1]
                doc_map[docid] = docno
            for docid in mq_set:
                print docid
                docno = doc_map[docid]
                query = r"update mq_text_label set DocNo = '%s' where DocID = '%s'"%(docno, docid)
                local_sql.excute(query)
                # booth_sql.excute(query)

        def update_numeric_docid(self):
                """
                update the Documents table to add numeric doc number
                """
                local_sql = MysqlConnecttion("local")
                booth_sql = MysqlConnecttion("booth")
                query = r"select docid from Documents"
                rows = local_sql.excute_with_result(query)
                count = 1
                for row in rows:
                        print count
                        query = r"update Documents set DocNo = '%d' where docid = '%s'"%(count, row[0])
                        count += 1
                        local_sql.excute(query)
                        booth_sql.excute(query)
        def fill_mq_with_no_mq_doc(self):
                """
                fill Magic_Quadrants table with Documents have no mqs in
                """
                local_sql = MysqlConnecttion("local")
                query = r"select distinct docid from Magic_Quadrants where removed = 0"
                rows = local_sql.excute_with_result(query)
                mq_doc_set = set()
                for row in rows:
                    mq_doc_set.add(row[0])
                query = r"select docid from Documents where MagicQuadrant = 1"
                rows = local_sql.excute_with_result(query)
                doc_set = set()
                for row in rows:
                    doc_set.add(row[0])
                mq_id = 10000
                for docid in doc_set:
                    if docid in mq_doc_set:
                        continue
                    query = r"insert into Magic_Quadrants (DocID, mqid, removed) values ('%s', 'mq%s', 0)"%(docid, mq_id)
                    mq_id += 1
                    local_sql.excute(query)
        def doc_title_get_word_vector(self):
            """
            extract the word vector from First_Paragraph, What_you_need_to_know and MarketDefinition
            """
            e =Extractor()
            local_sql = MysqlConnecttion("local")
            query = r"select distinct docid from doc_deatail_vector"
            rows = local_sql.excute_with_result(query)
            mq_set = set()
            for row in rows:
                mq_set.add(row[0])
            for docid in mq_set:
                print docid
                query = "select title from Documents WHERE docid = '%s'"%(docid)
                row = local_sql.excute_with_result(query)[0]
                content = row[0]
                if content == None or len(content) == 0:
                    continue
                word_vector = e.extract(content, shorten = False)
                try:
                        json_words = json.dumps(word_vector)
                        json_words = json_words.replace(r"'",r"\'")
                except UnicodeDecodeError, e:
                        print key
                        print pair_with_weight
                        json_words = ''
                        raw_input()
                query = r"UPDATE doc_deatail_vector set title = '%s' where docid = '%s'"%( json_words, docid)
                local_sql.excute(query)

        def doc_title_shorten_word_vector(self):
            """
            shorten the word vector from the vector we already got
            """
            field = "title"
            local_sql = MysqlConnecttion("local")
            doc_freq = set()
            query = r"select word from labels"
            rows = local_sql.excute_with_result(query)
            for row in rows:
                doc_freq.add(row[0])
            query = r'''SELECT DocID, %s FROM doc_deatail_vector'''%(field)
            rows = local_sql.excute_with_result(query)
            for row in rows:
                docid = row[0]
                json_words = row[1]
                if(json_words == None or len(json_words) == 0):
                    continue
                word_vector = json.loads(json_words)
                result = sorted(word_vector.items(), key=lambda d:(d[0].count("_")),reverse = True)
                short_vector = {}
                for tuple in result:
                        key = tuple[0]
                        value = tuple[1]
                        if(key in doc_freq):
                                is_in = 0
                                for word in short_vector:
                                        if(word.count(key) > 0):
                                                is_in += 1
                                if is_in == 0:
                                        short_vector[key] = value
                try:
                        if len(short_vector) > 0:
                                json_words = json.dumps(short_vector)
                                json_words = json_words.replace(r"'",r"\'")
                        else :
                                json_words = ""
                except UnicodeDecodeError, e:
                        print key
                        print short_vector
                        json_words = ''
                        raw_input()

                query = r"update doc_deatail_vector set %s_short = '%s' where DocID = '%s'"%(field, json_words, docid)
                local_sql.excute(query)
        def doc_title_mark_labels(self):
            """
            mark the doc title label
            """
            local_sql = MysqlConnecttion("local")
            # booth_sql = MysqlConnecttion("booth")
            label_map = {}#{"word":labelid}
            query = r"select Word, Labelid from labels WHERE Geogr = 0 and Vert = 0 and Mktseg = 0"
            rows = local_sql.excute_with_result(query)
            for row in rows:
                    label_map[row[0]] = row[1]
            doc_map = {}
            query = r'''SELECT docid, title_short from doc_deatail_vector
                                '''
            rows = local_sql.excute_with_result(query)
            for row in rows:
                docid = row[0]
                print docid
                title_label = set()
                json_words = row[1]
                word_vector = {}
                if(json_words == None or len(json_words) == 0):
                    pass
                else:
                    word_vector = json.loads(json_words)
                label_list = set()
                for word in word_vector:
                    if word in label_map:
                        if str(label_map[word]) not in label_list:
                            label_list.add(str(label_map[word]))
                title_labels = ";".join(label_list)
                query = r"update Magic_Quadrants set Labelid = '%s' where DocID = '%s' and removed = 0"%(title_labels, docid)
                # print query
                local_sql.excute(query)
        def doc_title_labels(self):
            self.doc_title_get_word_vector()
            self.doc_title_shorten_word_vector()
            self.doc_title_mark_labels()
        def repair(self):
            local_sql = MysqlConnecttion("local")
            query = r"select word, labelid from labels"
            rows = local_sql.excute_with_result(query)
            for row in rows:
                word = row[0]
                id = str(row[1])
                self.repair_label_id(word, id)
        def repair_label_id(self, word, id):
            local_sql = MysqlConnecttion("local")
            query = r'''SELECT docid, mqid, labelid FROM magic_quadrants
                                    WHERE mq_title_vector_short LIKE '%%%s%%'
                                    OR mq_title_vector_unique LIKE '%%%s%%'
                                '''%(word, word)
            rows = local_sql.excute_with_result(query)
            for row in rows:
                docid = row[0]
                mqid = row[1]
                labels = row[2]
                if labels == "":
                    labels = []
                else:
                    labels = labels.split(";")
                if id not in labels:
                    labels.append(id)
                labelid = ";".join(labels)
                query = r"update Magic_Quadrants set Labelid = '%s' where DocID = '%s' and mqid = '%s'"%(labelid, docid, mqid)
                print query
                local_sql.excute(query)

